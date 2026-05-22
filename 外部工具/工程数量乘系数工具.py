import argparse
import os
import re
import shutil
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl


HEADER_KEYWORDS = ("工程数量输入", "工程数量", "数量")
SKIP_HEADER_KEYWORDS = ("单价", "合价", "单重", "合重", "序号", "编号")
DEFAULT_HEADER_SCAN_ROWS = 20


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def parse_factor(text):
    text = str(text).strip()
    if not text:
        raise ValueError("请输入系数")
    try:
        value = Decimal(text)
    except InvalidOperation:
        raise ValueError("系数必须是数字，例如 1.15 或 0.8")
    if value == 0:
        raise ValueError("系数不能为 0")
    return text


def is_target_header(text):
    text = normalize_text(text)
    if not text:
        return False
    if any(keyword in text for keyword in SKIP_HEADER_KEYWORDS):
        return False
    return any(keyword in text for keyword in HEADER_KEYWORDS)


def find_target_columns(ws, header_row=None):
    rows = [header_row] if header_row else range(1, min(ws.max_row, DEFAULT_HEADER_SCAN_ROWS) + 1)
    found = []
    for row in rows:
        for col in range(1, ws.max_column + 1):
            header = normalize_text(ws.cell(row=row, column=col).value)
            if is_target_header(header):
                found.append((row, col, header))
        if found and not header_row:
            return found
    return found


def should_skip_cell(value):
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    if text.startswith("#"):
        return True
    return False


def multiply_expression(value, factor, repeat=False):
    if should_skip_cell(value):
        return None

    if isinstance(value, str):
        text = value.strip()
        if not repeat and re.search(r"\*[ ]*-?\d+(\.\d+)?$", text):
            return None
        if text.startswith("="):
            body = text[1:].strip()
            return f"=({body})*{factor}"
        if re.fullmatch(r"-?\d+(\.\d+)?", text):
            return f"{text}*{factor}"
        return f"({text})*{factor}"

    if isinstance(value, (int, float, Decimal)):
        return f"{value}*{factor}"

    return f"{value}*{factor}"


def make_backup(path):
    source = Path(path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = source.with_name(f"{source.stem}.bak_{stamp}{source.suffix}")
    shutil.copy2(source, backup)
    return backup


def process_workbook(path, factor, header_row=None, sheet_names=None, repeat=False, dry_run=False):
    suffix = Path(path).suffix.lower()
    if suffix != ".xlsx":
        raise ValueError("当前版本只支持 .xlsx 文件；请先在 Excel 中把 .xls 另存为 .xlsx 后再处理")

    wb = openpyxl.load_workbook(path)
    target_sheets = sheet_names or wb.sheetnames
    summary = []
    total_changed = 0

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"找不到工作表：{sheet_name}")

        ws = wb[sheet_name]
        columns = find_target_columns(ws, header_row)
        changed = 0
        touched_columns = []

        for header_row_index, col, header in columns:
            touched_columns.append(f"{header}({ws.cell(row=header_row_index, column=col).coordinate})")
            for row in range(header_row_index + 1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                new_value = multiply_expression(cell.value, factor, repeat=repeat)
                if new_value is not None:
                    changed += 1
                    total_changed += 1
                    if not dry_run:
                        cell.value = new_value

        summary.append((sheet_name, changed, touched_columns))

    backup = None
    if total_changed and not dry_run:
        backup = make_backup(path)
        wb.save(path)

    return {
        "total_changed": total_changed,
        "backup": str(backup) if backup else "",
        "summary": summary,
    }


def run_cli(args):
    factor = parse_factor(args.factor)
    sheet_names = [name.strip() for name in args.sheets.split(",") if name.strip()] if args.sheets else None
    result = process_workbook(
        args.file,
        factor,
        header_row=args.header_row,
        sheet_names=sheet_names,
        repeat=args.repeat,
        dry_run=args.dry_run,
    )
    print(format_result(result, dry_run=args.dry_run))


def format_result(result, dry_run=False):
    lines = []
    lines.append("预览完成。" if dry_run else "处理完成。")
    lines.append(f"改写单元格数量：{result['total_changed']}")
    if result.get("backup"):
        lines.append(f"原文件备份：{result['backup']}")
    lines.append("")
    for sheet_name, changed, columns in result["summary"]:
        cols = "、".join(columns) if columns else "未找到目标列"
        lines.append(f"{sheet_name}: {changed} 个；目标列：{cols}")
    return "\n".join(lines)


def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext

    root = tk.Tk()
    root.title("工程数量乘系数工具")
    root.geometry("760x520")
    root.minsize(720, 480)

    file_var = tk.StringVar()
    factor_var = tk.StringVar(value="1.0")
    header_row_var = tk.StringVar(value="")
    repeat_var = tk.BooleanVar(value=False)
    preview_var = tk.BooleanVar(value=True)

    def choose_file():
        filename = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
        )
        if filename:
            file_var.set(filename)

    def append_log(text):
        output.configure(state="normal")
        output.insert("end", text + "\n")
        output.see("end")
        output.configure(state="disabled")

    def execute():
        path = file_var.get().strip()
        if not path:
            messagebox.showwarning("缺少文件", "请先选择一个 .xlsx 文件")
            return
        if not os.path.exists(path):
            messagebox.showerror("文件不存在", "选择的文件不存在")
            return

        try:
            factor = parse_factor(factor_var.get())
            header_row = int(header_row_var.get()) if header_row_var.get().strip() else None
            result = process_workbook(
                path,
                factor,
                header_row=header_row,
                repeat=repeat_var.get(),
                dry_run=preview_var.get(),
            )
            append_log(format_result(result, dry_run=preview_var.get()))
            append_log("-" * 60)
            if preview_var.get():
                messagebox.showinfo("预览完成", "当前是预览模式，文件未被改动。取消勾选“只预览”后可正式处理。")
            else:
                messagebox.showinfo("处理完成", f"已处理 {result['total_changed']} 个单元格，并自动备份原文件。")
        except Exception as exc:
            messagebox.showerror("处理失败", str(exc))

    frame = tk.Frame(root, padx=14, pady=14)
    frame.pack(fill="both", expand=True)

    tk.Label(frame, text="Excel 文件").grid(row=0, column=0, sticky="w")
    tk.Entry(frame, textvariable=file_var).grid(row=0, column=1, sticky="ew", padx=8)
    tk.Button(frame, text="选择", command=choose_file, width=10).grid(row=0, column=2)

    tk.Label(frame, text="系数").grid(row=1, column=0, sticky="w", pady=(10, 0))
    tk.Entry(frame, textvariable=factor_var, width=18).grid(row=1, column=1, sticky="w", padx=8, pady=(10, 0))

    tk.Label(frame, text="标题行").grid(row=2, column=0, sticky="w", pady=(10, 0))
    tk.Entry(frame, textvariable=header_row_var, width=18).grid(row=2, column=1, sticky="w", padx=8, pady=(10, 0))

    tk.Checkbutton(frame, text="只预览，不改文件", variable=preview_var).grid(row=3, column=1, sticky="w", padx=4, pady=(12, 0))
    tk.Checkbutton(frame, text="允许对已乘系数的单元格再次处理", variable=repeat_var).grid(row=4, column=1, sticky="w", padx=4)

    tk.Button(frame, text="开始处理", command=execute, width=16, height=2).grid(row=3, column=2, rowspan=2, sticky="e")

    output = scrolledtext.ScrolledText(frame, height=18, state="disabled")
    output.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(16, 0))

    frame.columnconfigure(1, weight=1)
    frame.rowconfigure(5, weight=1)

    append_log("默认扫描前 20 行，寻找列标题：工程数量输入、工程数量、数量。")
    append_log("建议先用预览模式确认目标列，再正式处理。")
    root.mainloop()


def main():
    parser = argparse.ArgumentParser(description="批量把工程数量列改写为 原值*系数")
    parser.add_argument("file", nargs="?", help="要处理的 .xlsx 文件")
    parser.add_argument("factor", nargs="?", help="系数，例如 1.15 或 0.8")
    parser.add_argument("--header-row", type=int, default=None, help="指定标题行；不指定则扫描前 20 行")
    parser.add_argument("--sheets", default="", help="只处理指定工作表，多个用英文逗号分隔")
    parser.add_argument("--repeat", action="store_true", help="允许重复处理已经带 *系数 的单元格")
    parser.add_argument("--dry-run", action="store_true", help="只预览，不保存")
    args = parser.parse_args()

    if args.file and args.factor:
        run_cli(args)
    else:
        run_gui()


if __name__ == "__main__":
    main()
